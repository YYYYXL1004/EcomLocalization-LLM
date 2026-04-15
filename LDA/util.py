import numpy as np
import openpyxl
import pandas as pd
from openpyxl import load_workbook

# 编写函数对分类好的数据进行情感之分类计算
import NewExcel

#   国内装包镜像：pip install pandas -i https://pypi.tuna.tsinghua.edu.cn/simple




def comment_kind(kind_path):
    # 从excel文件中读取评论
    text = pd.read_excel('./excel/0.8.xlsx', usecols=['content'])
    label = pd.read_excel('./excel/0.8.xlsx', usecols=['label'])
    kind=pd.read_excel('./excel/0.8.xlsx', usecols=['label'])

    # 将DataFrame类型转化为list类型
    array = np.array(text)
    text_list = array.tolist()
    label_array = np.array(label)
    label_list = label_array.tolist()

    # print(text_list)
    # print(type(text_list))

    sentences = []
    emotion = []
    label_str = []

    for item in label_list:
        s1 = str(item)
        s2 = s1.replace('[', '')
        s3 = s2.replace(']', '')
        label_str.append(s3)



# 删除商品评论中特别短的评论
def update_data(path, file_name, sheetname):
    text = pd.read_excel(path, usecols=['content'])
    label = pd.read_excel(path, usecols=['label'])
    print(text)
    # row记录行号
    row=[]
    update_text=[]
    update_label=[]

    # 将DataFrame类型转化为list类型
    content_array = np.array(text)
    content_list = content_array.tolist()
    label_array = np.array(label)
    label_list = label_array.tolist()
    print(content_list)

    #print(content_list)
    for index,item in enumerate(content_list):
        if(len(item[0])>6):
            row.append(index)
            update_text.append(item[0])

            update_label.append(str(int(label_list[index][0])).replace('[','').replace(']',''))
    print(row)

    # 表头
    t = ['content', 'label']

    # 新建表格
    NewExcel.excel_int(file_name, sheetname)

    # 写入表头
    NewExcel.excel_write_title(file_name, t)

    # 写入三列数据
    NewExcel.excel_write_array_str(file_name, update_text, 0)
    NewExcel.excel_write_array_str(file_name, update_label, 1)

    print('删除短文本数据完成')




# 工具方法：excel文件 to txt文件
from spire.xls import *
from spire.xls.common import *
def ExcelToTxt():

    # 创建Workbook对象并加载Excel文件，文件的后缀可以是.xls或.xlsx
    workbook = Workbook()
    workbook.LoadFromFile('通讯录.xlsx')
    # 获取第一个工作表
    sheet = workbook.Worksheets[0]
    # 将工作表保存为TXT文本文件，以制表符（\t）作为数据分隔符(你也可以使用其他的符号作为数据分隔符，例如空格或逗号)
    sheet.SaveToFile('Excel转TXT.txt', '  ', Encoding.get_UTF8())
    workbook.Dispose()


# 工具方法：txt文件 to excel文件
def TxtToExcel():

    # 打开文本文件并逐行读取数据
    with open('销售数据.txt', 'r') as file:
        lines = file.readlines()
    # 将读取的每行文本数据进行处理，去除首尾空白字符并按制表符（\t）拆分为子字符串。将生成的子字符串存储在列表中
    data = [line.strip().split('\t') for line in lines]
    # 创建Workbook对象
    workbook = Workbook()
    # 获取第一个工作表
    sheet = workbook.Worksheets[0]
    # 遍历列表中的数据并将其填充到第一个工作表的单元格中
    for row_num, row_data in enumerate(data):
        for col_num, cell_data in enumerate(row_data):
            sheet.Range[row_num + 1, col_num + 1].Value = cell_data
            # 设置字体名称和大小
            sheet.Range[row_num + 1, col_num + 1].Style.Font.FontName = '宋体'
            sheet.Range[row_num + 1, col_num + 1].Style.Font.Size = 12
            # 将标题行的字体设置为粗体
            sheet.Range[1, col_num + 1].Style.Font.IsBold = True

    # 根据单元格内容自动调整表格列宽
    sheet.AllocatedRange.AutoFitColumns()

    # 保存结果为Excel XLSX文件
    workbook.SaveToFile('Text转Excel.xlsx', ExcelVersion.Version2016)
    # 或者保存结果为Excel XLS文件
    # workbook.SaveToFile('TextToExcel.xls', ExcelVersion.Version97to2003)
    workbook.Dispose()





if __name__ == "__main__":
    path='./excel/商品评论总和.xlsx'
    filename='./excel/商品评论总和.xlsx'
    update_data(path,filename,'测试数据')
