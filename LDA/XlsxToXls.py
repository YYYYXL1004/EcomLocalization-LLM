import os
import openpyxl
import win32com.client as win32
import pandas as pd
from openpyxl.utils import get_column_letter
def xlsx_to_xls(fname, export_name, delete_flag=True):
    """
    将xlsx文件转化为xls文件
    :param fname: 传入待转换的文件路径(可传绝对路径，也可传入相对路径，都可以)
    :param export_name: 传入转换后到哪个目录下的路径(可传绝对路径，也可传入相对路径，都可以)
    :param delete_flag: 转换成功后，是否删除原来的xlsx的文件,默认删除 布尔类型
    :return:    无返回值
    """
    excel = win32.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    absolute_path = os.path.join(os.path.dirname(os.path.abspath(fname)), os.path.basename(fname))
    save_path = os.path.join(os.path.dirname(os.path.abspath(export_name)), os.path.basename(export_name))
    wb = excel.Workbooks.Open(absolute_path)
    wb.SaveAs(save_path, FileFormat=56)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    if delete_flag:
        os.remove(absolute_path)

    print('转换完成')

#相对路径使用，可以转化





# 进行excel单元格内容的修改
def update_excel(file_path,sheet_name,update_list,col_name):

    # 根据列名获取列对应的数字
    df = pd.read_excel(file_path, usecols='D:F')
    offset = 4  # D

    col = get_column_letter(df.columns.get_loc(col_name) + offset)
    print(col)


    # 加载 excel 文件
    wb = openpyxl.load_workbook(file_path)

    # 得到sheet对象
    sheet = wb[sheet_name]
    for index,item in enumerate(update_list):
        #print(index,item)
        #print(type(item))
        #print()
        sheet[col+str(index+2)] = item
    ## 指定不同的文件名，可以另存为别的文件
    #wb.save('income-1.xlsx')
    wb.save(file_path)
    print('筛选词语完成')


# 读取txt文件
def file_read(path):  # 传入的参数为要获取文本的目标行数
    num = []  # 存档次数
    word_occur = []
    with open(path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        for item in lines:
            str_split=item.split(',')
            s1=str_split[0]+','+str_split[1]
            #print(s1,str_split[2])
            word_occur.append(s1)
            num.append(str_split[2])
        file.close()
    #print(word_occur)
    #print(num)
    print('读取edge.txt文件成功')
    return word_occur,num

# 将字典数据写入txt文件
def write_txt(path,list):
    Note = open(path, mode='w',encoding='utf-8')
    # 创建列表保存dic数据
    for index,item in enumerate(list):
        Note.write('Topic' + str(index) + ':\n')  # \n 换行符
        txt_list=[]
        for k in item:
            Note.write(k+item[k])
    Note.close()
    print('写入txt文件完成')


'''
if __name__ == '__main__':
    list1=[{'客服,购物': '26\n', '服务态度,购物': '9\n', '感觉,材料费': '5\n', '态度,电话': '21\n', '态度,时间': '7\n', '人员,时间': '4\n'}, {'信赖,品牌': '95\n', '信赖,质量': '25\n', '价格,质量': '22\n', '价格,实惠': '38\n', '材料,活动': '4\n'}]
    print(type(list1))
    for index,item in enumerate(list1):
        print(item)
    path=r'E:\python_code\LDA\csv\1.txt'
    write_txt(path,list1)
'''

